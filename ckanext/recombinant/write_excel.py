import textwrap

import openpyxl

from ckanext.recombinant.tables import get_geno
from ckanext.recombinant.errors import RecombinantException
from ckanext.recombinant.datatypes import datastore_type
from ckanext.recombinant.helpers import (
    recombinant_choice_fields, recombinant_language_text)

from ckan.plugins.toolkit import _

white_font = openpyxl.styles.Font(color=openpyxl.styles.colors.WHITE)

HEADER_ROW, HEADER_HEIGHT = 1, 27
CHEADINGS_ROW, CHEADINGS_HEIGHT = 2, 22
CHEADINGS_MIN_WIDTH = 10
LINE_HEIGHT = 15  # extra lines of text in same cell
CODE_ROW = 3
CSTATUS_ROW, CSTATUS_HEIGHT = 4, 6
EXAMPLE_ROW, EXAMPLE_HEIGHT = 5, 15
EXAMPLE_MERGE = 'A5:B5'
FREEZE_PANES = 'C5'
DATA_FIRST_ROW, DATA_HEIGHT = 6, 24
DATA_NUM_ROWS = 10000
RSTATUS_COL, RSTATUS_WIDTH = 'A', 1
RPAD_COL, RPAD_WIDTH = 'B', 2.5
DATA_FIRST_COL, DATA_FIRST_COL_NUM = 'C', 3
ESTIMATE_WIDTH_MULTIPLE = 1.2
EDGE_RANGE = 'A1:A4' # just to look spiffy

REF_HEADER1_ROW, REF_HEADER1_HEIGHT = 1, 27
REF_HEADER2_ROW, REF_HEADER2_HEIGHT = 2, 27
REF_FIRST_ROW = 4
REF_FIELD_NUM_COL_NUM = 1
REF_FIELD_NUM_MERGE = 'A{row}:B{row}'
REF_FIELD_TITLE_HEIGHT = 24
REF_FIELD_TITLE_MERGE = 'C{row}:D{row}'
REF_KEY_COL, REF_KEY_COL_NUM = 'C', 3
REF_KEY_WIDTH = 18
REF_VALUE_COL, REF_VALUE_COL_NUM = 'D', 4
REF_VALUE_WIDTH = 114
REF_EDGE_RANGE = 'A1:A2'

DEFAULT_EDGE_STYLE = {
    'PatternFill': {
        'patternType': 'solid',
        'fgColor': 'FF336B87'},
    'Font': {
        'color': 'FFFFFF'}}
DEFAULT_HEADER_STYLE = {
    'PatternFill': {
        'patternType': 'solid',
        'fgColor': 'FF90AFC5'},
    'Font': {
        'bold': True,
        'size': 16}}
DEFAULT_CHEADING_STYLE = {
    'PatternFill': {
        'patternType': 'solid',
        'fgColor': 'FF90AFC5'},
    'Font': {
        'underline': 'single'}}
DEFAULT_EXAMPLE_STYLE = {
    'PatternFill': {
        'patternType': 'solid',
        'fgColor': 'FFDDD9C4'}}
DEFAULT_ERROR_STYLE = {
    'PatternFill': {
        'patternType': 'solid',
        'fgColor': 'FFC00000'},
    'Font': {
        'color': 'FFFFFF'}}
DEFAULT_REF_HEADER2_STYLE = {
    'PatternFill': {
        'patternType': 'solid',
        'fgColor': 'FF90AFC5'},
    'Alignment': {
        'vertical': 'center'}}
REF_NUMBER_STYLE = {}
REF_TITLE_STYLE = {
    'PatternFill': {
        'patternType': 'solid',
        'fgColor': 'FFFFFFFF'},
    'Font': {
        'underline': 'single'}}
REF_ATTR_STYLE = {
    'PatternFill': {
        'patternType': 'solid',
        'fgColor': 'FFFFFFFF'},
    'Font': {
        'color': 'CCCCCC'},
    'Alignment': {
        'vertical': 'top'}}
REF_PAPER_STYLE = {
    'PatternFill': {
        'patternType': 'solid',
        'fgColor': 'FFFFFFFF'}}


def excel_template(dataset_type, org):
    """
    return an openpyxl.Workbook object containing the sheet and header fields
    for passed dataset_type and org. Supports version 2 and 3 templates.
    """
    geno = get_geno(dataset_type)
    version = geno.get('template_version', 2)

    book = openpyxl.Workbook()
    sheet = book.active
    refs = []
    for rnum, chromo in enumerate(geno['resources'], 1):
        if version == 2:
            _populate_excel_sheet_v2(sheet, chromo, org, refs)
        elif version == 3:
            _populate_excel_sheet(sheet, geno, chromo, org, refs, rnum)
        sheet = book.create_sheet()

    if version == 2:
        _populate_reference_sheet_v2(sheet, chromo, refs)
    elif version == 3:
        _populate_reference_sheet(sheet, chromo, refs)
    sheet.title = 'reference'

    if version == 2:
        return book

    for i, chromo in enumerate(geno['resources']):
        sheet = book.create_sheet()
        #_populate_excel_validation(sheet, chromo, org, refs)
    return book


def excel_data_dictionary(geno):
    """
    return an openpyxl.Workbook object containing the field reference
    from geno, one sheet per language
    """
    book = openpyxl.Workbook()
    sheet = book.active

    style1 = {
        'PatternFill': {
            'patternType': 'solid',
            'fgColor': 'FFFFF056'},
        'Font': {
            'bold': True}}
    style2 = {
        'PatternFill': {
            'patternType': 'solid',
            'fgColor': 'FFDFE2DB'}}

    from pylons import config
    from ckan.lib.i18n import handle_request, get_lang
    from ckan.common import c, request

    for lang in config['ckan.locales_offered'].split():
        if sheet is None:
            sheet = book.create_sheet()

        sheet.title = lang.upper()
        # switch language (FIXME: this is harder than it should be)
        request.environ['CKAN_LANG'] = lang
        handle_request(request, c)
        choice_fields = dict(
            (f['datastore_id'], f['choices'])
            for f in recombinant_choice_fields(chromo['resource_name'])
            for chromo in geno['resources'])

        refs = []
        for chromo in geno['resources']:
            for field in chromo['fields']:
                _append_field_ref_rows(refs, field, style1, style2)

                if field['datastore_id'] in choice_fields:
                    _append_field_choices_rows(
                        refs,
                        choice_fields[field['datastore_id']],
                        style2)

        _populate_reference_sheet(sheet, geno, refs)
        sheet = None

    return book


def estimate_width(text):
    return max(len(s) for s in text.split('\n')) * ESTIMATE_WIDTH_MULTIPLE

def wrap_text_to_width(text, width):
    cwidth = width // ESTIMATE_WIDTH_MULTIPLE
    return '\n'.join(
        '\n'.join(textwrap.wrap(line, cwidth))
        for line in text.split('\n'))


def _populate_excel_sheet(sheet, geno, chromo, org, refs, resource_num):
    """
    Format openpyxl sheet for the resource definition chromo and org.
    (Version 3)

    refs - list of rows to add to reference sheet, modified
        in place from this function
    resource_num - 1-based index of resource
    """
    sheet.title = chromo['resource_name']

    edge_style = geno.get('excel_edge_style', DEFAULT_EDGE_STYLE)
    required_style = geno.get('excel_required_style', edge_style)
    header_style = geno.get('excel_header_style', DEFAULT_HEADER_STYLE)
    cheadings_style = geno.get('excel_column_heading_style', DEFAULT_CHEADING_STYLE)
    example_style = geno.get('excel_example_style', DEFAULT_EXAMPLE_STYLE)
    errors_style = geno.get('excel_error_style', DEFAULT_ERROR_STYLE)

    # create rows so we can set all heights
    for i in xrange(1, DATA_FIRST_ROW + DATA_NUM_ROWS):
        sheet.append([])

    sheet.merge_cells(EXAMPLE_MERGE)
    fill_cell(sheet, EXAMPLE_ROW, 1, _('e.g.'), example_style)

    fill_cell(
        sheet,
        HEADER_ROW,
        DATA_FIRST_COL_NUM,
        recombinant_language_text(chromo['title']) + '        ' + org['title'],
        header_style)

    sheet.cell(row=CODE_ROW, column=1).value = 'v3'  # template version
    sheet.cell(row=CODE_ROW, column=2).value = org['name']  # allow only upload to this org

    cheadings_dimensions = sheet.row_dimensions[CHEADINGS_ROW]

    choice_fields = dict(
        (f['datastore_id'], f['choices'])
        for f in recombinant_choice_fields(chromo['resource_name']))

    for col_num, field in enumerate(
            (f for f in chromo['fields'] if f.get(
                'import_template_include', True)), DATA_FIRST_COL_NUM):
        field_heading = recombinant_language_text(
            field.get('excel_heading', field['label'])).strip()
        cheadings_dimensions.height = max(
            cheadings_dimensions.height,
            field_heading.count('\n') * LINE_HEIGHT + CHEADINGS_HEIGHT)
        fill_cell(
            sheet,
            CHEADINGS_ROW,
            col_num,
            hyperlink_formula(
                '#reference!{col}{row}'.format(
                    col=REF_KEY_COL,
                    row=len(refs) + REF_FIRST_ROW),
                field_heading),
            cheadings_style)

        # match against db columns
        sheet.cell(row=CODE_ROW, column=col_num).value = field['datastore_id']

        example = chromo['examples']['record'].get(field['datastore_id'])
        if example:
            fill_cell(
                sheet,
                EXAMPLE_ROW,
                col_num,
                u','.join(example) if isinstance(example, list)
                else unicode(example),
                example_style)

        col_letter = openpyxl.cell.get_column_letter(col_num)
        col_letter_before = openpyxl.cell.get_column_letter(max(1, col_num-1))
        col_letter_after = openpyxl.cell.get_column_letter(col_num+1)

        col = sheet.column_dimensions[col_letter]
        if 'excel_column_width' in field:
            col.width = field['excel_column_width']
        else:
            col.width = max(estimate_width(field_heading), CHEADINGS_MIN_WIDTH)

        validation_range = '{col}{row1}:{col}{rowN}'.format(
            col=col_letter,
            row1=DATA_FIRST_ROW,
            rowN=DATA_FIRST_ROW + DATA_NUM_ROWS - 1)

        xl_format = datastore_type[field['datastore_type']].xl_format
        alignment = openpyxl.styles.Alignment(wrap_text=True)
        protection = openpyxl.styles.Protection(locked=False)
        for (c,) in sheet[validation_range]:
            c.number_format = xl_format
            c.alignment = alignment
            c.protection = protection
        ex_cell = sheet.cell(row=EXAMPLE_ROW, column=col_num)
        ex_cell.number_format = xl_format
        ex_cell.alignment = alignment

        _append_field_ref_rows(refs, field, '#{sheet}!{col}{row}'.format(
            sheet=sheet.title, col=col_letter, row=CHEADINGS_ROW))

        if field['datastore_id'] in choice_fields:
            ref1 = len(refs) + REF_FIRST_ROW
            _append_field_choices_rows(
                refs,
                choice_fields[field['datastore_id']])
            refN = len(refs) + REF_FIRST_ROW - 2

            choice_range = 'reference!${col}${ref1}:${col}${refN}'.format(
                col=REF_KEY_COL, ref1=ref1, refN=refN)

            choices = [c[0] for c in choice_fields[field['datastore_id']]]
            if field['datastore_type'] != '_text':
                v = openpyxl.worksheet.datavalidation.DataValidation(
                    type="list",
                    formula1=choice_range,
                    allow_blank=True)
                v.errorTitle = u'Invalid choice'
                valid_keys = u', '.join(unicode(c) for c in choices)
                if len(valid_keys) < 40:
                    v.error = (u'Please enter one of the valid keys: '
                        + valid_keys)
                else:
                    v.error = (u'Please enter one of the valid keys shown on '
                        'sheet "reference" rows {0}-{1}'.format(ref1, refN))
                sheet.add_data_validation(v)
                v.ranges.append(validation_range)

    sheet.row_dimensions[HEADER_ROW].height = HEADER_HEIGHT
    sheet.row_dimensions[CODE_ROW].hidden = True
    sheet.row_dimensions[CSTATUS_ROW].height = CSTATUS_HEIGHT
    sheet.row_dimensions[EXAMPLE_ROW].height = EXAMPLE_HEIGHT
    for i in xrange(DATA_FIRST_ROW, DATA_FIRST_ROW + DATA_NUM_ROWS):
        sheet.row_dimensions[i].height = DATA_HEIGHT

    sheet.column_dimensions[RSTATUS_COL].width = RSTATUS_WIDTH
    sheet.column_dimensions[RPAD_COL].width = RPAD_WIDTH

    sheet.freeze_panes = sheet[FREEZE_PANES]

    apply_styles(header_style, sheet.row_dimensions[HEADER_ROW])
    apply_styles(cheadings_style, sheet.row_dimensions[CHEADINGS_ROW])
    apply_styles(cheadings_style, sheet.row_dimensions[CSTATUS_ROW])
    apply_styles(example_style, sheet.row_dimensions[EXAMPLE_ROW])
    for (c,) in sheet[EDGE_RANGE]:
        apply_styles(edge_style, c)

    select = "{col}{row}".format(col=DATA_FIRST_COL, row=DATA_FIRST_ROW)
    sheet.sheet_view.selection[0].activeCell = select
    sheet.sheet_view.selection[0].sqref = select



def _populate_excel_sheet_v2(sheet, chromo, org, refs):
    """
    Format openpyxl sheet for the resource definition chromo and org.

    refs - list of rows to add to reference sheet, modified
        in place from this function

    returns field information for reference sheet
    """
    boolean_validator = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1='"FALSE,TRUE"', allow_blank=True)
    sheet.add_data_validation(boolean_validator)

    sheet.title = chromo['resource_name']

    org_style = dict(
        chromo['excel_organization_style'],
        Alignment={'vertical': 'center'})
    fill_cell(sheet, 1, 1, org['name'], org_style)
    fill_cell(
        sheet,
        1,
        2,
        recombinant_language_text(chromo['title']) + '        ' + org['title'],
        org_style)
    sheet.row_dimensions[1].height = 24
    apply_styles(org_style, sheet.row_dimensions[1])

    header_style = chromo['excel_header_style']
    error_color = chromo.get('excel_error_background_color', '763626')
    required_color = chromo.get('excel_required_border_color', '763626')

    error_fill = openpyxl.styles.PatternFill(
        start_color='FF%s' % error_color,
        end_color='FF%s' % error_color,
        fill_type='solid')
    required_side = openpyxl.styles.Side(
        style='medium',
        color='FF%s' % required_color)
    required_border = openpyxl.styles.Border(
        required_side, required_side, required_side, required_side)


    choice_fields = dict(
        (f['datastore_id'], f['choices'])
        for f in recombinant_choice_fields(chromo['resource_name']))

    pk_cells = [
        openpyxl.cell.get_column_letter(n)+'4' for
        n, field in enumerate((f for f in chromo['fields'] if f.get(
                    'import_template_include', True)), 1)
        if field['datastore_id'] in chromo['datastore_primary_key']]

    for n, field in enumerate((f for f in chromo['fields'] if f.get(
            'import_template_include', True)), 1):
        fill_cell(sheet, 2, n, recombinant_language_text(field['label']), header_style)
        fill_cell(sheet, 3, n, field['datastore_id'], header_style)
        # jumping through openpyxl hoops:
        col_letter = openpyxl.cell.get_column_letter(n)
        col_letter_before = openpyxl.cell.get_column_letter(max(1, n-1))
        col_letter_after = openpyxl.cell.get_column_letter(n+1)
        col = sheet.column_dimensions[col_letter]
        col.width = field['excel_column_width']
        col.alignment = openpyxl.styles.Alignment(
            wrap_text=True)
        # FIXME: format only below header
        col.number_format = datastore_type[field['datastore_type']].xl_format
        validation_range = '{0}4:{0}1004'.format(col_letter)

        _append_field_ref_rows_v2(refs, field, org_style, header_style)

        if field['datastore_type'] == 'boolean':
            boolean_validator.ranges.append(validation_range)
        if field['datastore_type'] == 'date':
            sheet.conditional_formatting.add(validation_range,
                openpyxl.formatting.FormulaRule([
                        # +0 is needed by excel to recognize dates. sometimes.
                        'AND(NOT(ISBLANK({cell})),NOT(ISNUMBER({cell}+0)))'
                        .format(cell=col_letter + '4',)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
            sheet.conditional_formatting.add("{0}2".format(col_letter),
                openpyxl.formatting.FormulaRule([
                        # +0 is needed by excel to recognize dates. sometimes.
                        'SUMPRODUCT(--NOT(ISBLANK({cells})),'
                        '--NOT(ISNUMBER({cells}+0)))'
                        .format(cells=validation_range,)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
        if field['datastore_type'] == 'int':
            sheet.conditional_formatting.add(validation_range,
                openpyxl.formatting.FormulaRule([
                        'AND(NOT(ISBLANK({cell})),NOT(IFERROR(INT({cell})={cell},FALSE)))'
                        .format(cell=col_letter + '4',)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
            sheet.conditional_formatting.add("{0}2".format(col_letter),
                openpyxl.formatting.FormulaRule([
                        'SUMPRODUCT(--NOT(ISBLANK({cells})),'
                        '--NOT(IFERROR(INT({cells})={cells},FALSE)))'
                        .format(cells=validation_range,)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
        if field['datastore_type'] == 'money':
            sheet.conditional_formatting.add(validation_range,
                openpyxl.formatting.FormulaRule([
                        # isblank doesnt work. sometimes. trim()="" is more permissive
                        'AND(NOT(TRIM({cell})=""),NOT(IFERROR(ROUND({cell},2)={cell},FALSE)))'
                        .format(cell=col_letter + '4',)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
            sheet.conditional_formatting.add("{0}2".format(col_letter),
                openpyxl.formatting.FormulaRule([
                        # isblank doesnt work. sometimes. trim()="" is more permissive
                        'SUMPRODUCT(--NOT(TRIM({cells})=""),'
                        '--NOT(IFERROR(ROUND({cells},2)={cells},FALSE)))'
                        .format(cells=validation_range,)],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))


        if field['datastore_id'] in choice_fields:
            ref1 = len(refs) + 1
            _append_field_choices_rows_v2(
                refs,
                choice_fields[field['datastore_id']],
                header_style,
                sheet.title + '!' + validation_range
                if field['datastore_type'] == '_text' else None)
            refN = len(refs)

            choice_range = 'reference!$B${0}:$B${1}'.format(ref1, refN)

            choices = [c[0] for c in choice_fields[field['datastore_id']]]
            if field['datastore_type'] == '_text':
                # custom validation only works in Excel, use conditional
                # formatting for libre office compatibility
                sheet.conditional_formatting.add(validation_range,
                    openpyxl.formatting.FormulaRule([(
                        # count characters in the cell
                        'IF(SUBSTITUTE({col}4," ","")="",0,'
                        'LEN(SUBSTITUTE({col}4," ",""))+1)-'
                        # minus length of valid choices
                        'SUMPRODUCT(--ISNUMBER(SEARCH('
                        '","&{r}&",",SUBSTITUTE(","&{col}4&","," ",""))),'
                        'LEN({r})+1)'
                        .format(
                            col=col_letter,
                            r=choice_range)
                        )],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))
            else:
                v = openpyxl.worksheet.datavalidation.DataValidation(
                    type="list",
                    formula1=choice_range,
                    allow_blank=True)
                v.errorTitle = u'Invalid choice'
                valid_keys = u', '.join(unicode(c) for c in choices)
                if len(valid_keys) < 40:
                    v.error = (u'Please enter one of the valid keys: '
                        + valid_keys)
                else:
                    v.error = (u'Please enter one of the valid keys shown on '
                        'sheet "reference" rows {0}-{1}'.format(ref1, refN))
                sheet.add_data_validation(v)
                v.ranges.append(validation_range)

            # hilight header if bad values pasted below
            if field['datastore_type'] == '_text':
                choice_counts = 'reference!$J${0}:$J${1}'.format(ref1, refN)
                sheet.conditional_formatting.add("{0}2".format(col_letter),
                    openpyxl.formatting.FormulaRule([(
                            # count characters in the validation range
                            'SUMPRODUCT(IF(SUBSTITUTE({v}," ","")="",0,'
                            'LEN(SUBSTITUTE({v}," ",""))+1))-'
                            # minus length of all valid choices found
                            'SUMPRODUCT({counts},LEN({choices})+1)'
                            .format(
                                v=validation_range,
                                col=col_letter,
                                choices=choice_range,
                                counts=choice_counts)
                            )],
                        stopIfTrue=True,
                        fill=error_fill,
                        font=white_font,
                        ))
            else:
                sheet.conditional_formatting.add("{0}2".format(col_letter),
                    openpyxl.formatting.FormulaRule([(
                            'SUMPRODUCT(--NOT(TRIM({0})=""))'
                            '-SUMPRODUCT(COUNTIF({1},TRIM({0})))'
                            .format(validation_range, choice_range))],
                        stopIfTrue=True,
                        fill=error_fill,
                        font=white_font,
                        ))

        if field.get('excel_cell_required_formula'):
            sheet.conditional_formatting.add(validation_range,
                openpyxl.formatting.FormulaRule([
                        field['excel_cell_required_formula'].format(
                            column=col_letter,
                            column_before=col_letter_before,
                            column_after=col_letter_after,
                            row='4',
                        )],
                    stopIfTrue=True,
                    border=required_border,
                    ))
        elif (field.get('excel_required') or
                field['datastore_id'] in chromo['datastore_primary_key']):
            # hilight missing values
            sheet.conditional_formatting.add(validation_range,
                openpyxl.formatting.FormulaRule([(
                        'AND({col}4="",SUMPRODUCT(LEN(A4:Z4)))'
                        .format(col=col_letter)
                        )],
                    stopIfTrue=True,
                    border=required_border,
                    ))
        if field.get('excel_cell_error_formula'):
            sheet.conditional_formatting.add(validation_range,
                openpyxl.formatting.FormulaRule([
                    field['excel_cell_error_formula'].format(
                        cell=col_letter + '4',)
                    ],
                stopIfTrue=True,
                fill=error_fill,
                font=white_font,
                ))
        if field.get('excel_header_error_formula'):
            sheet.conditional_formatting.add("{0}2".format(col_letter),
                openpyxl.formatting.FormulaRule([
                        field['excel_header_error_formula'].format(
                            cells=validation_range,
                            column=col_letter,
                        )],
                    stopIfTrue=True,
                    fill=error_fill,
                    font=white_font,
                    ))

    apply_styles(header_style, sheet.row_dimensions[2])
    apply_styles(header_style, sheet.row_dimensions[3])
    sheet.row_dimensions[3].hidden = True

    sheet.freeze_panes = sheet['A4']


def hyperlink_formula(target, text):
    text = text.replace(u'"', u'""')
    text = text.replace(u'\n', u'"&CHAR(10)&"')
    return u'=HYPERLINK("{target}","{text}")'.format(target=target, text=text)


def _append_field_ref_rows(refs, field, link):
    refs.append((None, []))
    refs.append(('title', [
        hyperlink_formula(
            link,
            recombinant_language_text(field['label']))]))
    refs.append(('attr', [
        _('ID'),
        field['datastore_id']]))
    if 'description' in field:
        refs.append(('attr', [
            _('Description'),
            recombinant_language_text(field['description'])]))
    if 'obligation' in field:
        refs.append(('attr', [
            _('Obligation'),
            recombinant_language_text(field['obligation'])]))
    if 'format_type' in field:
        refs.append(('attr', [
            _('Format'),
            recombinant_language_text(field['format_type'])]))

def _append_field_choices_rows(refs, choices):
    refs.append(('attr', [_('Values')]))
    for key, value in choices:
        if unicode(key) != value:
            refs.append(('choice', [unicode(key), value]))
        else:
            refs.append(('choice', [unicode(key)]))


def _append_field_ref_rows_v2(refs, field, style1, style2):
    a1 = (style2, style1, 24)
    a2 = (style2, None, None)
    refs.append((None, []))
    refs.append((a1, [
        _('Field Name'),
        recombinant_language_text(field['label'])]))
    refs.append((a2, [
        _('ID'),
        field['datastore_id']]))
    if 'description' in field:
        refs.append((a2, [
            _('Description'),
            recombinant_language_text(field['description'])]))
    if 'obligation' in field:
        refs.append((a2, [
            _('Obligation'),
            recombinant_language_text(field['obligation'])]))
    if 'format_type' in field:
        refs.append((a2, [
            _('Format'),
            recombinant_language_text(field['format_type'])]))

def _append_field_choices_rows_v2(refs, choices, style2, count_range=None):
    label = _('Values')
    a1 = (style2, None, 24)
    for key, value in choices:
        r = [label, unicode(key), value]
        if count_range: # used by _text choices validation
            r.extend([None]*6 + ['=SUMPRODUCT(--ISNUMBER(SEARCH('
                '","&B{n}&",",SUBSTITUTE(","&{r}&","," ",""))))'.format(
                    r=count_range,
                    n=len(refs) + 1)])
        refs.append((a1, r))
        label = None
        a1 = (style2, None, None)

def _populate_reference_sheet(sheet, geno, refs):
    field_count = 1

    edge_style = geno.get('excel_edge_style', DEFAULT_EDGE_STYLE)
    header1_style = geno.get('excel_header_style', DEFAULT_HEADER_STYLE)
    header2_style = geno.get('excel_ref_header2_style', DEFAULT_REF_HEADER2_STYLE)
    choice_style = geno.get('excel_example_style', DEFAULT_EXAMPLE_STYLE)

    fill_cell(
        sheet,
        REF_HEADER1_ROW,
        REF_KEY_COL_NUM,
        recombinant_language_text(geno['title']),
        header1_style)
    apply_styles(header1_style, sheet.row_dimensions[REF_HEADER1_ROW])
    fill_cell(
        sheet,
        REF_HEADER2_ROW,
        REF_KEY_COL_NUM,
        _('Reference'),
        header2_style)
    apply_styles(header2_style, sheet.row_dimensions[REF_HEADER2_ROW])
    for (c,) in sheet[REF_EDGE_RANGE]:
        apply_styles(edge_style, c)
    sheet.row_dimensions[REF_HEADER1_ROW].height = REF_HEADER1_HEIGHT
    sheet.row_dimensions[REF_HEADER2_ROW].height = REF_HEADER2_HEIGHT


    for row_number, (style, ref_line) in enumerate(refs, 3):
        if len(ref_line) == 2:
            value = wrap_text_to_width(ref_line[1], REF_VALUE_WIDTH)
            ref_line = [ref_line[0], value]

        sheet.append([None, None] + ref_line)

        if len(ref_line) == 2:
            sheet.row_dimensions[row_number].height = LINE_HEIGHT + (
                value.count('\n') * LINE_HEIGHT)

        if style == 'title':
            sheet.merge_cells(REF_FIELD_NUM_MERGE.format(row=row_number))
            sheet.merge_cells(REF_FIELD_TITLE_MERGE.format(row=row_number))
            fill_cell(
                sheet,
                row_number,
                REF_FIELD_NUM_COL_NUM,
                '{fnum}.'.format(fnum=field_count),
                REF_NUMBER_STYLE)
            title_cell = sheet.cell(row=row_number, column=REF_KEY_COL_NUM)
            apply_styles(REF_TITLE_STYLE, title_cell)
            sheet.row_dimensions[row_number].height = REF_FIELD_TITLE_HEIGHT
            field_count += 1

        elif style == 'choice':
            pad_cell = sheet.cell(row=row_number, column=REF_KEY_COL_NUM - 1)
            key_cell = sheet.cell(row=row_number, column=REF_KEY_COL_NUM)
            value_cell = sheet.cell(row=row_number, column=REF_VALUE_COL_NUM)
            apply_styles(choice_style, pad_cell)
            apply_styles(choice_style, key_cell)
            apply_styles(choice_style, value_cell)

        elif style == 'attr':
            key_cell = sheet.cell(row=row_number, column=REF_KEY_COL_NUM)
            apply_styles(REF_ATTR_STYLE, key_cell)

        apply_styles(REF_PAPER_STYLE, sheet.row_dimensions[row_number])

    sheet.column_dimensions[RSTATUS_COL].width = RSTATUS_WIDTH
    sheet.column_dimensions[RPAD_COL].width = RPAD_WIDTH
    sheet.column_dimensions[REF_KEY_COL].width = REF_KEY_WIDTH
    sheet.column_dimensions[REF_VALUE_COL].width = REF_VALUE_WIDTH


def _populate_reference_sheet_v2(sheet, chromo, refs):
    for style, ref_line in refs:
        sheet.append(ref_line)
        if not style:
            continue

        s1, s2, height = style
        if height:
            sheet.row_dimensions[sheet.max_row].height = height

        if s2:
            apply_styles(s2, sheet.row_dimensions[sheet.max_row])
        for c in range(len(ref_line)):
            if c and s2:
                apply_styles(s2, sheet.cell(
                    row=sheet.max_row, column=c + 1))
            if not c and s1:
                apply_styles(s1, sheet.cell(
                    row=sheet.max_row, column=c + 1))


def fill_cell(sheet, row, column, value, styles):
    c = sheet.cell(row=row, column=column)
    c.value = value.replace('\n', '\r\n')
    apply_styles(styles, c)


def apply_styles(config, target):
    """
    apply styles from config to target

    currently supports PatternFill, Font, Alignment
    """
    pattern_fill = config.get('PatternFill')
    if pattern_fill:
        target.fill = openpyxl.styles.PatternFill(**pattern_fill)
    font = config.get('Font')
    if font:
        target.font = openpyxl.styles.Font(**font)
    alignment = config.get('Alignment')
    if alignment:
        target.alignment = openpyxl.styles.Alignment(**alignment)
